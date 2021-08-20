import pandas
import openpyxl
import time
from datetime import date, datetime
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from os import path

import utils

BUDGET_ENTRIES = [
    "Staff Regional Meetings",
    "Staff Training",
    "Staff Expenses",
    "Revenue Related",
    "IT - Related",
    "Occupancy - Related",
    "Transport & Travels",
    "Other Related",
    "Allocated Expenses"
]

def parse_excel_to_budgets(file):
    """Returns a tuple of :
    - Two list of dict (MonthlyActual, MonthlyBudget) 
    - One dict (YearlyBudget)"""
    res_actual = []
    res_budget = []
    res_yearlyBudget = {}

    folder_name, file_name = path.split(file)

    # Data Validation
    if not (path.exists(file) and folder_name and file_name):
        raise FileNotFoundError("Given filepath doesn't exist!")
    
    # Slimify xlsx
    slim_filepath = _slimify_mrpt(folder_name, file_name)
    # slim_filepath = "data/MRPT dummy_slim.xlsx"
    wb = openpyxl.load_workbook(filename=slim_filepath)
    ws = wb.active

    # Get Cells
    (mActualCells, mBudgetCells, yBudgetCell) = _get_mrpt_cells(ws)

    # Get Entry Labels Column
    entryLabels = _find_entry_labels_columns(ws)
    
    # Process Actuals
    for a in mActualCells:
        res_actual.append(_process_column(ws, a, entryLabels))
    
    # Process mBudgets
    for b in mBudgetCells:
        res_budget.append(_process_column(ws, b, entryLabels))

    # Process yBudget
    res_yearlyBudget = _process_column(ws, yBudgetCell, entryLabels)

    return (res_actual, res_budget, res_yearlyBudget)

def _get_mrpt_cells(ws: Worksheet):
    top_header_cell = _find_cell_by_value(ws, "MTD Actual")
    
    if not top_header_cell:
        raise TypeError("Unknown Excel data structure!")
    
    # Validate Data
    months_of_actual = _count_cells_by_value(ws, "MTD Actual")
    months_of_budget = _count_cells_by_value(ws, "MTD Budget")

    if not (months_of_actual == months_of_budget):
        raise TypeError("Different amount of MonthlyActual and MonthlyBudget in file!")
    if not (_count_cells_by_value(ws,"YTD Budget") == 2):
        raise TypeError("There aren't two columns of YTD Budget!")

    # Get Cells
    actuals = _find_headercells_by_value(ws, "MTD Actual", months_of_actual)
    budgets = _find_headercells_by_value(ws, "MTD Budget", months_of_budget)
    yearlyBudget = _find_yearlybudget_cell(ws, "YTD Budget")

    return (actuals, budgets, yearlyBudget)

def _slimify_mrpt(dirName, fileName):
    try:
        df = pandas.read_excel(path.join(dirName, fileName), sheet_name="MRPT")
    except ValueError:
        return ValueError("Fail to read excel using Pandas")

    new_filename = fileName.split('.')[0] + "_slim.xlsx"

    new_path = path.join(dirName, new_filename)

    df.to_excel(new_path, index=False, header=True)

    return new_path

def _find_cell_by_value(ws: Worksheet, val: str):
    for row in ws.rows:
        for cell in row:
            if cell.value == val:
                return cell
    return None

def _find_headercells_by_value(ws: Worksheet, val: str, limit:int):
    """Returns of dict of (header (Cell) and date (datetime) of Headercell"""
    res = []

    found = 0

    for row in ws.rows:
        for cell in row:
            if cell.value == val:
                # Check if the monthyear's syntax below the found cell is correct 
                month_cell = cell.offset(row=1, column=0)
                try:
                    monthyear_data = datetime.strptime(month_cell.value, "%b %Y")
                except ValueError:
                    raise TypeError(f"Wrong Month Year format ({month_cell.value}) found on Cell {month_cell.coordinate}!")

                res.append({
                    "header": cell,
                    "date": monthyear_data
                })

                found += 1

                if found == limit:
                    return res
    
    raise TypeError(f"Only found {found} of {limit} '{val}' cells")

def _find_yearlybudget_cell(ws: Worksheet, val: str):
    for row in ws.rows:
        for cell in row:
            if cell.value == val:
                month_cell = cell.offset(row=1, column=0)
                try:
                    monthyear_data = datetime.strptime(month_cell.value, "%b %Y")
                except ValueError:
                    raise TypeError(f"Wrong Month Year format ({month_cell.value}) found on Cell {month_cell.coordinate}!")

                if monthyear_data.month == 12:
                    return {
                        "header": cell,
                        "date": monthyear_data
                    }
    raise TypeError(f"No ({val}) was found!")

def _count_cells_by_value(ws: Worksheet, val: str):
    count = 0

    for row in ws.rows:
        for cell in row:
            if cell.value == val:
                count += 1
    
    return count

def _find_entry_labels_columns(ws: Worksheet):
    """Returns a list of dict ('entry' (str: entry name) and 'cell' (Cell) of entry labels"""
    res = []

    for e in BUDGET_ENTRIES:
        res.append({
            'entry': e,
            'cell': None
        })

    # Find all entry label cells
    for row in ws:
        for cell in row:
            if cell.value in BUDGET_ENTRIES:
                index = utils.find_index(res, 'entry', cell.value)
                res[index]['cell'] = cell

    # Check if all entry cells are found
    for r in res:
        if r['cell'] == None:
            raise TypeError(f"Missing entry label ({r['entry']})")

    return res

def _process_column(ws: Worksheet, header_cell, entryLabels):
    r = {}

    column_index = header_cell['header'].column

    r["year"]      = header_cell['date'].year
    r["month"]     = header_cell['date'].month

    staff_total_expense = ws.cell(row=_get_label_cell_row(entryLabels, "Staff Expenses"), column=column_index).value
    staff_training      = ws.cell(row=_get_label_cell_row(entryLabels, "Staff Training"), column=column_index).value
    staff_regmeet       = ws.cell(row=_get_label_cell_row(entryLabels, "Staff Regional Meetings"), column=column_index).value
    
    staff_train_regmeet = staff_training + staff_regmeet
    staff_salaries      = staff_total_expense - staff_train_regmeet

    other_total             = ws.cell(row=_get_label_cell_row(entryLabels, "Other Related"), column=column_index).value
    other_transport_travel  = ws.cell(row=_get_label_cell_row(entryLabels, "Transport & Travels"), column=column_index).value
    
    other_other             = other_total - other_transport_travel


    r["staff_salaries"]             = staff_salaries
    r["staff_training_reg_meeting"] = staff_train_regmeet
    r["revenue_related"]            = ws.cell(row=_get_label_cell_row(entryLabels, "Revenue Related"), column=column_index).value
    r["it_related"]                 = ws.cell(row=_get_label_cell_row(entryLabels, "IT - Related"), column=column_index).value
    r["occupancy_related"]          = ws.cell(row=_get_label_cell_row(entryLabels, "Occupancy - Related"), column=column_index).value
    r["other_transport_travel"]     = other_transport_travel
    r["other_other"]                = other_other
    r["indirect_expense"]           = ws.cell(row=_get_label_cell_row(entryLabels, "Allocated Expenses"), column=column_index).value

    return r

def _get_label_cell_row(entryLabels, value):
    return entryLabels[utils.find_index(entryLabels, "entry", value)]["cell"].row

# foldername = "data" 
# filename = "MRPT dummy.xlsx"

# start_time = time.time()
# a, b, y = parse_excel_to_budgets(path.join(foldername, filename))
# print(f"Normal takes {time.time() - start_time}")